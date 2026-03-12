from selenium.webdriver import Chrome

import json
from openpyxl import load_workbook
from pathlib import Path
ROOT_PATH=Path(__file__).parent.parent
class BasePage:

    def __init__(self):
        self.url='https://www.dell.com/support/home/en-in'

    def support_login(self):
        url=self.url
        # with open('../testdata/user_data.json','r') as user_json:
        #     user_info=json.load(user_json)
        # username=user_info['user']["support_user_password"]
        # print(username)

        #reading from excel
        wbook=load_workbook(r'E:\Tekpyramid\python_selenium_automation\selenium_learning\Framework_sriram\testdata\users.xlsx')
        wsheet=wbook.active

        username=wsheet.cell(2,1).value
        pwd=wsheet.cell(2,2).value
        print(username,pwd)

         #sign in option
        from selenium import webdriver

        o=webdriver.ChromeOptions()
        o.add_experimental_option('detach',True)

        d=webdriver.Chrome(o)

        d.get(self.url)
        from time import sleep
        from selenium.webdriver.common.action_chains import ActionChains

        sleep(5)
        action=ActionChains(d)
        signin=d.find_element('xpath',"//span[text()='Sign In']/ancestor::a")
        # signin.click()
        action.move_to_element(signin).perform()
        sleep(5)
        print(d.current_url)
        signin_button=d.find_element('css selector','div.mh-myaccount-ctas>a')
        signin_button.click()
        print(d.current_url)
        d.find_element('id',"SignInModel_EmailAddress").send_keys(username)
        sleep(20)
        continue_button=d.find_element("id","btnContinue")
        continue_button.click()
        print(d.current_url)
        sleep(10)

        password_field=d.find_element("id","userPwd_UserInputSecret")
        password_field.send_keys(pwd)
        sleep(5)
        final_signin_button = d.find_element("id", "btnSignIn")
        final_signin_button.click()
        print(d.current_url)
        sleep(25)
        print(d.current_url)

# BasePage().support_login()