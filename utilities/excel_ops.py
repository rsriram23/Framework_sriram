from openpyxl import load_workbook
from pathlib import Path


ROOT_PATH=Path(__file__).parent.parent

class ExcelOperations:

    def get_data_from_users_excel(self,type_of_cred):
        # reading from excel
        wbook = load_workbook(fr'{ROOT_PATH}\testdata\users.xlsx')
        wsheet = wbook.active
        sheet_rows=list(wsheet.iter_rows(values_only=True))
        credential_combination={}
        for a_row in sheet_rows:
            credential_combination[a_row[2]]=(a_row[0],a_row[1])
        # username = wsheet.cell(2, 1).value
        # pwd = wsheet.cell(2, 2).value
        credentials_needed=credential_combination[type_of_cred]
        print(credentials_needed) #(username, pwd)
        return credentials_needed #username, pwd
